import streamlit as st
import pandas as pd
from sqlalchemy import create_engine, text
from datetime import datetime, timedelta, date, timezone
import calendar
import io
import json
import zipfile
from weasyprint import HTML
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- 1. Database Connection ---
try:
    if "DB_URL" not in st.secrets:
        st.error("❌ `DB_URL` missing in Streamlit Secrets! Please configure it in Settings -> Secrets.")
        st.stop()
    
    DB_URL = st.secrets["DB_URL"]
    engine = create_engine(DB_URL)
    with engine.connect() as conn: pass
except Exception as db_err:
    st.error("### 🛑 Database Connection Critical Failure")
    st.info(f"**Error Details:**\n`{str(db_err)}`")
    st.stop()

# --- 2. Utility Functions & Dynamic Years (HKT Timezone locked) ---
HKT = timezone(timedelta(hours=8))

def to_date(val):
    try:
        if pd.isna(val) or val == "" or str(val).strip() == "" or str(val).lower() in ["none", "nat", "nan"]: return None
        return pd.to_datetime(val).date()
    except: return None

def clean_val(v):
    v = str(v).strip()
    if v.lower() in ["nat", "none", "nan", ""]: return ""
    if v.endswith(" 00:00:00"): return v.replace(" 00:00:00", "")
    return v

def get_anniv(year, month, day):
    try: return date(year, month, day)
    except ValueError: return date(year, month, day - 1)

def clean_status(val):
    v = str(val)
    for emo in ["🔴 ", "🟡 ", "🟢 ", "✅ ", "⚪ "]: v = v.replace(emo, "")
    return v

def add_one_month(dt):
    if not dt: return None
    month = dt.month + 1
    year = dt.year
    if month > 12:
        month = 1
        year += 1
    day = min(dt.day, calendar.monthrange(year, month)[1])
    return date(year, month, day)

def get_base_date(row_dict):
    place = str(row_dict.get('incorp_place', ''))
    is_hk_reg = str(row_dict.get('is_hk_registered', 'False')).strip().lower() in ['true', 'yes', 'y', '1']
    if place == 'HK': return to_date(row_dict.get('incorp_date'))
    if is_hk_reg: return to_date(row_dict.get('hk_incorp_date'))
    return None

current_system_year = datetime.now(HKT).year
active_years = list(range(2025, current_system_year + 1))

# --- 3. Navigation ---
st.set_page_config(page_title="ERP Cloud V166", layout="wide")
choice = st.sidebar.radio("Navigation", ["📊 Dashboard", "🏢 Company Register", "⚙️ Group Management", "📤 Data Exchange"])

TEMPLATE_COLS = [
    "client_group", "name_en", "name_ch", "incorp_place", "incorp_place_others", 
    "incorp_date", "ci_no", "is_hk_registered", "hk_incorp_date", "hk_ci_no", "br_no", 
    "co_type", "reg_addr", "corres_addr", "round_loc", "sign_loc", "seal_loc", 
    "nd2a_eff_date", "nd2a_file_date", "nd2a_download", "nd4_eff_date", "nd4_file_date", "nd4_download", 
    "nn6_eff_date", "nn6_file_date", "nn6_download",
    "dissolution_date", "remark"
]

# --- 4. Report Generation (Added Cache to Prevent Timeout) ---
@st.cache_data(show_spinner=False)
def generate_custom_pdf(selected_df, hide_client_group=False):
    now = datetime.now(HKT).strftime("%Y/%m/%d %H:%M")
    def fmt_date(val):
        d = to_date(val)
        return d.strftime('%Y/%m/%d') if d else "N/A"
    
    if not selected_df.empty:
        sort_cols = [c for c in ['client_group', 'name_en', 'incorp_place'] if c in selected_df.columns]
        selected_df = selected_df.sort_values(by=sort_cols, na_position='last')

    html_header = """
    <html>
    <head>
        <meta charset="UTF-8">
        <link href="https://fonts.googleapis.com/css2?family=Noto+Sans+TC:wght@400;700&display=swap" rel="stylesheet">
        <style>
            @page { size: A4; margin: 15mm 10mm 25mm 10mm; }
            body { font-family: 'Noto Sans TC', sans-serif; color: #2c3e50; line-height: 1.4; background-color: #ffffff; text-align: justify; }
            #footer { position: fixed; bottom: -15mm; left: 0; right: 0; width: 100%; border-top: 1px solid #eee; padding-top: 5px; font-size: 9pt; color: #7f8c8d; }
            .footer-table { width: 100%; border-collapse: collapse; }
            .footer-left { text-align: left; width: 50%; }
            .footer-right { text-align: right; width: 50%; }
            .company-container { page-break-before: always; width: 100%; }
            .company-container:first-child { page-break-before: auto; }
            .main-table { width: 100%; border-collapse: collapse; }
            .main-table thead { display: table-header-group; }
            .header-content { text-align: center; border-bottom: 1px solid #eee; padding-bottom: 15px; margin-bottom: 20px; }
            .name-en { font-size: 20pt; font-weight: bold; color: #2980b9; text-align: center; }
            .name-ch { font-size: 15pt; color: #333333; margin-top: 5px; text-align: center; min-height: 20px; }
            .section-bar { background-color: #f1f4f6; padding: 8px 15px; font-weight: bold; font-size: 11pt; margin: 20px 0 10px 0; border-left: 5px solid #3498db; color: #2c3e50; text-align: left; }
            .section-group { page-break-inside: avoid; }
            .info-table { width: 100%; border-collapse: collapse; }
            .info-table tr { border-bottom: 1px solid #f1f2f6; page-break-inside: avoid; }
            .info-table th { text-align: left; width: 45%; color: #7f8c8d; padding: 8px 0; font-weight: normal; font-size: 10.5pt; }
            .info-table td { text-align: justify; padding: 8px 0; color: #2c3e50; font-size: 10.5pt; font-weight: bold; }
        </style>
    </head>
    <body>
        <div id="footer">
            <table class="footer-table">
                <tr>
                    <td class="footer-left">Corporate Portfolio Report</td>
                    <td class="footer-right">Generated on: """ + now + """</td>
                </tr>
            </table>
        </div>
    """

    card_template = """
    <div class="company-container">
        <table class="main-table">
            <thead>
                <tr><td><div class="header-content"><div class="name-en">__NAME_EN__</div><div class="name-ch">__NAME_CH__</div></div></td></tr>
            </thead>
            <tbody>
                <tr>
                    <td>
                        <div class="company-card">
                            <div class="section-group">
                                <div class="section-bar">Registration Details</div>
                                <table class="info-table">
                                    __CLIENT_GROUP_ROW__
                                    <tr><th>Incorp. Place</th><td>__INCORP_PLACE__</td></tr>
                                    __DYNAMIC_PLACE_ROWS__
                                    __DYNAMIC_HK_ROWS__
                                    <tr><th>Company Type</th><td>__CO_TYPE__</td></tr>
                                </table>
                            </div>
                            <div class="section-group">
                                <div class="section-bar">Addresses</div>
                                <table class="info-table">
                                    <tr><th>Registered Address</th><td>__REG_ADDR__</td></tr>
                                    <tr><th>Correspondence Address</th><td>__CORRES_ADDR__</td></tr>
                                </table>
                            </div>
                            <div class="section-group">
                                <div class="section-bar">Items Storage</div>
                                <table class="info-table">
                                    <tr><th>Round Stamp</th><td>__ROUND_LOC__</td></tr>
                                    <tr><th>Signature Chop</th><td>__SIGN_LOC__</td></tr>
                                    <tr><th>Common Seal</th><td>__SEAL_LOC__</td></tr>
                                </table>
                            </div>
                            <div class="section-group">
                                <div class="section-bar">Compliance Filings (Yearly)</div>
                                <table class="info-table">
                                    __DYNAMIC_ANNUAL_ROWS__
                                </table>
                            </div>
                            __DYNAMIC_SEC_ROWS__
                            <div class="section-group">
                                <div class="section-bar">Remarks</div>
                                <div style="padding: 8px 15px; font-size: 10.5pt; color: #2c3e50; white-space: pre-wrap;">__REMARK__</div>
                            </div>
                        </div>
                    </td>
                </tr>
            </tbody>
        </table>
    </div>
    """

    final_html = html_header
    cg_row_html = "" if hide_client_group else "<tr><th>Client Group</th><td>__CLIENT_GROUP__</td></tr>"
    
    for _, row in selected_df.iterrows():
        ch_name = row.get('name_ch', '')
        if not ch_name or pd.isna(ch_name): ch_name = ''
        
        place = str(row.get('incorp_place', ''))
        is_hk_reg = str(row.get('is_hk_registered', 'False')) == 'True'
        base_date = get_base_date(row)
        incorp_year = base_date.year if base_date else None
        
        dynamic_place_rows = ""
        display_place = place
        if place == 'Others': display_place = f"Others ({str(row.get('incorp_place_others', ''))})"
            
        dynamic_place_rows += f"<tr><th>{place} Incorp. Date (YYYY/MM/DD)</th><td>{fmt_date(row.get('incorp_date'))}</td></tr>"
        dynamic_place_rows += f"<tr><th>{place} CI No.</th><td>{str(row.get('ci_no', ''))}</td></tr>"

        dynamic_hk_rows = ""
        dynamic_annual_rows = ""
        
        if place == 'HK' or is_hk_reg:
            if place == 'HK':
                dynamic_hk_rows += f"<tr><th>HK BR No.</th><td>{str(row.get('br_no', ''))}</td></tr>"
            else:
                dynamic_hk_rows += f"<tr><th>HK Incorp. Date (YYYY/MM/DD)</th><td>{fmt_date(row.get('hk_incorp_date'))}</td></tr>"
                dynamic_hk_rows += f"<tr><th>HK CI No.</th><td>{str(row.get('hk_ci_no', ''))}</td></tr>"
                dynamic_hk_rows += f"<tr><th>HK BR No.</th><td>{str(row.get('br_no', ''))}</td></tr>"

            comp_rec_str = str(row.get('compliance_records', '{}'))
            try: rec_dict = json.loads(comp_rec_str)
            except: rec_dict = {}
            if not isinstance(rec_dict, dict): rec_dict = {}
            
            for y in active_years:
                if incorp_year and y < incorp_year:
                    dynamic_annual_rows += f"<tr><th colspan='2' style='background-color:#f8f9fa; color:#2980b9; text-align:center;'>--- {y} Annual Cycle ---</th></tr>"
                    dynamic_annual_rows += f"<tr><td colspan='2' style='text-align:center; color:#7f8c8d; font-weight:normal;'>Not Incorporated Yet</td></tr>"
                else:
                    y_str = str(y)
                    y_data = rec_dict.get(y_str, {})
                    br_by = y_data.get('br_paid_by', 'Firm')
                    br_dt = y_data.get('br_date', 'N/A')
                    ar_dt = y_data.get('ar_date', 'N/A')
                    
                    if incorp_year and y == incorp_year and (not ar_dt or ar_dt == 'N/A'):
                        ar_dt_disp = "Exempt (1st Year)"
                    else:
                        ar_dt_disp = ar_dt if ar_dt else 'N/A'
                        
                    if br_dt and br_dt != 'N/A': br_dt = br_dt.replace('-', '/')
                    if ar_dt_disp and ar_dt_disp not in ['N/A', 'Exempt (1st Year)']: ar_dt_disp = ar_dt_disp.replace('-', '/')
                        
                    dynamic_annual_rows += f"<tr><th colspan='2' style='background-color:#f8f9fa; color:#2980b9; text-align:center;'>--- {y} Annual Cycle ---</th></tr>"
                    dynamic_annual_rows += f"<tr><th>BR Paid By ({y})</th><td>{br_by}</td></tr>"
                    dynamic_annual_rows += f"<tr><th>BR Paid Date ({y}) (YYYY/MM/DD)</th><td>{br_dt if br_dt else 'N/A'}</td></tr>"
                    dynamic_annual_rows += f"<tr><th>AR Filed Date ({y}) (YYYY/MM/DD)</th><td>{ar_dt_disp}</td></tr>"
        
        dynamic_sec_rows = ""
        if place == 'HK':
            dynamic_sec_rows += f"""
            <div class="section-group">
                <div class="section-bar">Company Secretary Actions</div>
                <table class="info-table">
                    <tr><th>ND2A Eff. Date (YYYY/MM/DD)</th><td>{fmt_date(row.get('nd2a_eff_date'))}</td></tr>
                    <tr><th>ND4 Eff. Date (YYYY/MM/DD)</th><td>{fmt_date(row.get('nd4_eff_date'))}</td></tr>
                </table>
            </div>"""
        elif is_hk_reg:
            dynamic_sec_rows += f"""
            <div class="section-group">
                <div class="section-bar">Non-HK Company Secretary Actions</div>
                <table class="info-table">
                    <tr><th>NN6 Eff. Date (YYYY/MM/DD)</th><td>{fmt_date(row.get('nn6_eff_date'))}</td></tr>
                </table>
            </div>"""

        remark_val = str(row.get('remark', ''))
        if remark_val == 'None' or not remark_val: remark_val = 'No remarks.'

        card = card_template.replace("__CLIENT_GROUP_ROW__", cg_row_html)
        card = card.replace("__NAME_EN__", str(row.get('name_en', '')))
        card = card.replace("__NAME_CH__", str(ch_name))
        card = card.replace("__CLIENT_GROUP__", str(row.get('client_group', '')))
        card = card.replace("__INCORP_PLACE__", display_place)
        card = card.replace("__DYNAMIC_PLACE_ROWS__", dynamic_place_rows)
        card = card.replace("__DYNAMIC_HK_ROWS__", dynamic_hk_rows)
        card = card.replace("__CO_TYPE__", str(row.get('co_type', '')))
        card = card.replace("__REG_ADDR__", str(row.get('reg_addr', '')))
        card = card.replace("__CORRES_ADDR__", str(row.get('corres_addr', '')))
        card = card.replace("__ROUND_LOC__", str(row.get('round_loc', '')))
        card = card.replace("__SIGN_LOC__", str(row.get('sign_loc', '')))
        card = card.replace("__SEAL_LOC__", str(row.get('seal_loc', '')))
        card = card.replace("__DYNAMIC_ANNUAL_ROWS__", dynamic_annual_rows)
        card = card.replace("__DYNAMIC_SEC_ROWS__", dynamic_sec_rows)
        card = card.replace("__REMARK__", remark_val)
        final_html += card

    final_html += "</body></html>"
    return HTML(string=final_html).write_pdf()

@st.cache_data(show_spinner=False)
def generate_general_excel(selected_df, hide_client_group=False):
    buf = io.BytesIO()
    df_export = selected_df.copy()
    
    for y in active_years:
        df_export[f'{y} BR Paid By'] = 'Firm'
        df_export[f'{y} BR Date'] = ''
        df_export[f'{y} AR Date'] = ''
        
    for idx, row in df_export.iterrows():
        base_date = get_base_date(row)
        incorp_year = base_date.year if base_date else None
        
        comp_rec_str = str(row.get('compliance_records', '{}'))
        try: rec_dict = json.loads(comp_rec_str)
        except: rec_dict = {}
        if isinstance(rec_dict, dict):
            for y in active_years:
                if incorp_year and y < incorp_year:
                    df_export.at[idx, f'{y} BR Paid By'] = 'N/A'
                    continue
                    
                y_str = str(y)
                y_data = rec_dict.get(y_str, {})
                df_export.at[idx, f'{y} BR Paid By'] = y_data.get('br_paid_by', 'Firm')
                
                br_d = y_data.get('br_date', '')
                ar_d = y_data.get('ar_date', '')
                if br_d: br_d = br_d.replace('-', '/')
                if ar_d: ar_d = ar_d.replace('-', '/')
                
                df_export.at[idx, f'{y} BR Date'] = br_d
                df_export.at[idx, f'{y} AR Date'] = ar_d
                
    base_cols = [c for c in TEMPLATE_COLS if c in df_export.columns and c != 'remark']
    
    dyn_cols = []
    for y in active_years:
        dyn_cols.extend([f"{y} BR Paid By", f"{y} BR Date", f"{y} AR Date"])
    
    df_export = df_export[base_cols + dyn_cols + ['remark']] 
    date_cols = ["incorp_date", "hk_incorp_date", "nd2a_eff_date", "nd2a_file_date", "nd4_eff_date", "nd4_file_date", "nn6_eff_date", "nn6_file_date", "dissolution_date"]
    for col in date_cols:
        if col in df_export.columns: df_export[col] = pd.to_datetime(df_export[col], errors='coerce').dt.strftime('%Y/%m/%d')
            
    if hide_client_group and 'client_group' in df_export.columns:
        df_export = df_export.drop(columns=['client_group'])
        
    df_export.to_excel(buf, index=False)
    return buf.getvalue()

@st.cache_data(show_spinner=False)
def generate_beautiful_excel(df, hide_client_group=False):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Outstanding Compliance"
    ws.views.sheetView[0].showGridLines = True

    font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="Arial", size=10)
    font_red = Font(name="Arial", size=10, color="FF0000", bold=True)
    font_yellow = Font(name="Arial", size=10, color="FF9900", bold=True)
    font_green = Font(name="Arial", size=10, color="00B050", bold=True)
    font_grey = Font(name="Arial", size=10, color="7F8C8D", italic=True)

    fill_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    fill_zebra = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    thin_border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))

    headers = ["Client Group", "Company Name EN", "Company Name CH", "Incorp Place", "Year", "BR Paid By", "BR Status", "BR Deadline (YYYY/MM/DD)", "AR Status", "AR Deadline (YYYY/MM/DD)", "Remark"]
    if hide_client_group: headers.remove("Client Group")

    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    ws["A1"] = f"Outstanding Compliance Report (Generated on: {datetime.now(HKT).strftime('%Y/%m/%d %H:%M')})"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color="1F497D")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[3].height = 25

    center_cols = ["Client Group", "Incorp Place", "Year", "BR Paid By", "BR Deadline (YYYY/MM/DD)", "AR Deadline (YYYY/MM/DD)"]

    current_row = 4
    for _, item in df.iterrows():
        row_dict = {
            "Client Group": item.get("Client Group", ""),
            "Company Name EN": item.get("Company Name EN", ""),
            "Company Name CH": item.get("Company Name CH", ""),
            "Incorp Place": item.get("Incorp Place", ""),
            "Year": str(item.get("Year", "")),
            "BR Paid By": item.get("BR Paid By", ""),
            "BR Status": item.get("BR Status", ""),
            "BR Deadline (YYYY/MM/DD)": item.get("BR Deadline", ""),
            "AR Status": item.get("AR Status", ""),
            "AR Deadline (YYYY/MM/DD)": item.get("AR Deadline", ""),
            "Remark": item.get("Remark", "")
        }
        
        for col_idx, h in enumerate(headers, 1):
            val_str = clean_status(row_dict[h]) if row_dict[h] else ""
            cell = ws.cell(row=current_row, column=col_idx, value=val_str)
            cell.font = font_data
            cell.border = thin_border
            if current_row % 2 == 0: cell.fill = fill_zebra
            
            if h in center_cols: cell.alignment = Alignment(horizontal="center", vertical="center")
            else: cell.alignment = Alignment(horizontal="left", vertical="center")
                
            if h in ["BR Status", "AR Status"]:
                if "Overdue" in val_str: cell.font = font_red
                elif "Due Soon" in val_str: cell.font = font_yellow
                elif "Exempt" in val_str or "Not Incorporated" in val_str or "Pending" in val_str: cell.font = font_grey
                else: cell.font = font_green
                    
        ws.row_dimensions[current_row].height = 20
        current_row += 1

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1: continue 
            if cell.value: max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

@st.cache_data(show_spinner=False)
def generate_outstanding_pdf(df, hide_client_group=False):
    now_str = datetime.now(HKT).strftime("%Y/%m/%d %H:%M")
    colspan = "10" if hide_client_group else "11"
    cg_th = "" if hide_client_group else '<th style="width:8%">Client Group</th>'
    
    html = f"""
    <html>
    <head>
        <meta charset="UTF-8">
        <link href="https://fonts.googleapis.com/css2?family=Noto+Sans+TC:wght@400;700&display=swap" rel="stylesheet">
        <style>
            @page {{ size: A4 landscape; margin: 15mm; background-color: #ffffff; }}
            body {{ font-family: 'Noto Sans TC', sans-serif; font-size: 9pt; color: #2c3e50; margin: 0; padding: 0; }}
            table {{ width: 100%; border-collapse: collapse; }}
            thead {{ display: table-header-group; }}
            tr {{ page-break-inside: avoid; }}
            th {{ background-color: #1f497d; color: white; padding: 6px 4px; border: 1px solid #d9d9d9; font-size: 9pt; text-align: center; vertical-align: middle; }}
            td {{ padding: 6px 4px; border: 1px solid #d9d9d9; text-align: center; vertical-align: middle; }}
            tr:nth-child(even) td {{ background-color: #f8f9fa; }}
            .text-left {{ text-align: left; font-weight: bold; color: #2980b9; }}
            .report-header-cell {{ border: none !important; background-color: white !important; text-align: left !important; padding: 0 0 15px 0 !important; }}
        </style>
    </head>
    <body>
        <table>
            <thead>
                <tr>
                    <td colspan="{colspan}" class="report-header-cell">
                        <h2 style="color: #1f497d; margin: 0 0 5px 0;">Outstanding Compliance Report</h2>
                        <p style="color: #7f8c8d; font-size: 9pt; margin: 0;">Generated on: {now_str}</p>
                    </td>
                </tr>
                <tr>
                    {cg_th}
                    <th style="width:16%">Company Name EN</th>
                    <th style="width:12%">Company Name CH</th>
                    <th style="width:8%">Incorp Place</th>
                    <th style="width:5%">Year</th>
                    <th style="width:7%">BR Paid By</th>
                    <th style="width:8%">BR Status</th>
                    <th style="width:8%">BR Deadline<br>(YYYY/MM/DD)</th>
                    <th style="width:8%">AR Status</th>
                    <th style="width:8%">AR Deadline<br>(YYYY/MM/DD)</th>
                    <th style="width:12%">Remark</th>
                </tr>
            </thead>
            <tbody>
    """
    for _, r in df.iterrows():
        br_val = clean_status(r.get('BR Status', ''))
        ar_val = clean_status(r.get('AR Status', ''))
        
        def get_color(v):
            if "Overdue" in v: return "#ff0000"
            if "Due Soon" in v: return "#ff9900"
            if "Exempt" in v or "Not Incorporated" in v or "Pending" in v: return "#7f8c8d"
            return "#00b050"
            
        br_color = get_color(br_val)
        ar_color = get_color(ar_val)
        cg_td = "" if hide_client_group else f"<td>{r.get('Client Group', '')}</td>"
        
        html += f"""
        <tr>
            {cg_td}
            <td class="text-left">{r.get('Company Name EN', '')}</td>
            <td class="text-left">{r.get('Company Name CH', '')}</td>
            <td>{r.get('Incorp Place', '')}</td>
            <td style="font-weight: bold; color: #1f497d;">{r.get('Year', '')}</td>
            <td>{r.get('BR Paid By', '')}</td>
            <td style="color: {br_color}; font-weight: bold;">{br_val}</td>
            <td>{r.get('BR Deadline', '')}</td>
            <td style="color: {ar_color}; font-weight: bold;">{ar_val}</td>
            <td>{r.get('AR Deadline', '')}</td>
            <td style="text-align: left; font-size: 8.5pt; color: #7f8c8d;">{r.get('Remark', '')}</td>
        </tr>
        """
    html += "</tbody></table></body></html>"
    return HTML(string=html).write_pdf()

@st.cache_data(show_spinner=False)
def generate_inv_excel(df, hide_client_group=False):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoicing Schedule"
    ws.views.sheetView[0].showGridLines = True

    font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    font_data = Font(name="Arial", size=10)
    fill_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    fill_zebra = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    thin_border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))

    headers = ["Client Group", "Company Name EN", "Company Name CH", "Incorp Place", "Year", "Anniversary (MM/DD)", "BR Paid By", "BR Deadline (YYYY/MM/DD)", "AR Deadline (YYYY/MM/DD)", "Remark"]
    if hide_client_group: headers.remove("Client Group")

    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    ws["A1"] = f"Invoicing Schedule Report (Generated on: {datetime.now(HKT).strftime('%Y/%m/%d %H:%M')})"
    ws["A1"].font = Font(name="Arial", size=14, bold=True, color="1F497D")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=text)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[3].height = 25

    center_cols = ["Client Group", "Incorp Place", "Year", "Anniversary (MM/DD)", "BR Paid By", "BR Deadline (YYYY/MM/DD)", "AR Deadline (YYYY/MM/DD)"]

    current_row = 4
    for _, item in df.iterrows():
        row_dict = {
            "Client Group": item.get("Client Group", ""),
            "Company Name EN": item.get("Company Name EN", ""),
            "Company Name CH": item.get("Company Name CH", ""),
            "Incorp Place": item.get("Incorp Place", ""),
            "Year": str(item.get("Year", "")),
            "Anniversary (MM/DD)": item.get("Anniversary (MM/DD)", ""),
            "BR Paid By": item.get("BR Paid By", ""),
            "BR Deadline (YYYY/MM/DD)": item.get("BR Deadline", ""),
            "AR Deadline (YYYY/MM/DD)": item.get("AR Deadline", ""),
            "Remark": item.get("Remark", "")
        }
        
        for col_idx, h in enumerate(headers, 1):
            val = row_dict[h]
            cell = ws.cell(row=current_row, column=col_idx, value=val if val else "")
            cell.font = font_data
            cell.border = thin_border
            if current_row % 2 == 0: cell.fill = fill_zebra
            
            if h in center_cols: cell.alignment = Alignment(horizontal="center", vertical="center")
            else: cell.alignment = Alignment(horizontal="left", vertical="center")
                    
        ws.row_dimensions[current_row].height = 20
        current_row += 1

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1: continue 
            if cell.value: max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

@st.cache_data(show_spinner=False)
def generate_inv_pdf(df, year, month_disp, hide_client_group=False):
    now_str = datetime.now(HKT).strftime("%Y/%m/%d %H:%M")
    colspan = "9" if hide_client_group else "10"
    cg_th = "" if hide_client_group else '<th style="width:9%">Client Group</th>'
    
    html = f"""
    <html>
    <head>
        <meta charset="UTF-8">
        <link href="https://fonts.googleapis.com/css2?family=Noto+Sans+TC:wght@400;700&display=swap" rel="stylesheet">
        <style>
            @page {{ size: A4 landscape; margin: 15mm; background-color: #ffffff; }}
            body {{ font-family: 'Noto Sans TC', sans-serif; font-size: 9pt; color: #2c3e50; margin: 0; padding: 0; }}
            table {{ width: 100%; border-collapse: collapse; }}
            thead {{ display: table-header-group; }}
            tr {{ page-break-inside: avoid; }}
            th {{ background-color: #1f497d; color: white; padding: 6px 4px; border: 1px solid #d9d9d9; font-size: 9pt; text-align: center; vertical-align: middle; }}
            td {{ padding: 6px 4px; border: 1px solid #d9d9d9; text-align: center; vertical-align: middle; }}
            tr:nth-child(even) td {{ background-color: #f8f9fa; }}
            .text-left {{ text-align: left; font-weight: bold; color: #2980b9; }}
            .report-header-cell {{ border: none !important; background-color: white !important; text-align: left !important; padding: 0 0 15px 0 !important; }}
        </style>
    </head>
    <body>
        <table>
            <thead>
                <tr>
                    <td colspan="{colspan}" class="report-header-cell">
                        <h2 style="color: #1f497d; margin: 0 0 5px 0;">Invoicing Schedule Report ({year} - Months: {month_disp})</h2>
                        <p style="color: #7f8c8d; font-size: 9pt; margin: 0;">Generated on: {now_str}</p>
                    </td>
                </tr>
                <tr>
                    {cg_th}
                    <th style="width:18%">Company Name EN</th>
                    <th style="width:13%">Company Name CH</th>
                    <th style="width:9%">Incorp Place</th>
                    <th style="width:5%">Year</th>
                    <th style="width:8%">Anniversary<br>(MM/DD)</th>
                    <th style="width:8%">BR Paid By</th>
                    <th style="width:9%">BR Deadline<br>(YYYY/MM/DD)</th>
                    <th style="width:9%">AR Deadline<br>(YYYY/MM/DD)</th>
                    <th style="width:12%">Remark</th>
                </tr>
            </thead>
            <tbody>
    """
    for _, r in df.iterrows():
        cg_td = "" if hide_client_group else f"<td>{r.get('Client Group', '')}</td>"
        html += f"""
        <tr>
            {cg_td}
            <td class="text-left">{r.get('Company Name EN', '')}</td>
            <td class="text-left">{r.get('Company Name CH', '')}</td>
            <td>{r.get('Incorp Place', '')}</td>
            <td style="font-weight: bold; color: #1f497d;">{r.get('Year', '')}</td>
            <td style="font-weight: bold;">{r.get('Anniversary (MM/DD)', '')}</td>
            <td>{r.get('BR Paid By', '')}</td>
            <td>{r.get('BR Deadline', '')}</td>
            <td>{r.get('AR Deadline', '')}</td>
            <td style="text-align: left; font-size: 8.5pt; color: #7f8c8d;">{r.get('Remark', '')}</td>
        </tr>
        """
    html += "</tbody></table></body></html>"
    return HTML(string=html).write_pdf()

# --- ZIP Export Logic (Cached & Rewritten for Stability) ---
@st.cache_data(show_spinner=False)
def create_zip_pdfs(df, report_type="All", year=None, month_disp=None):
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", compression=zipfile.ZIP_DEFLATED) as zip_file:
        col_name = 'client_group' if 'client_group' in df.columns else 'Client Group'
        groups = df[col_name].unique()
        for g in groups:
            group_df = df[df[col_name] == g]
            safe_g = str(g).replace("/", "_").replace("\\", "_")
            if not safe_g.strip(): safe_g = "Ungrouped"
            
            if report_type == "All":
                pdf_bytes = generate_custom_pdf(group_df, hide_client_group=True)
                filename = f"{safe_g}_Company_Report.pdf"
            elif report_type == "Outstanding":
                pdf_bytes = generate_outstanding_pdf(group_df, hide_client_group=True)
                filename = f"{safe_g}_Outstanding_Compliance.pdf"
            elif report_type == "Invoicing":
                pdf_bytes = generate_inv_pdf(group_df, year, month_disp, hide_client_group=True)
                filename = f"{safe_g}_Invoicing_Schedule_{year}_{month_disp}.pdf"
            
            zip_file.writestr(filename, pdf_bytes)
    return zip_buffer.getvalue()

@st.cache_data(show_spinner=False)
def create_zip_excels(df, report_type="All", year=None, month_disp=None):
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", compression=zipfile.ZIP_DEFLATED) as zip_file:
        col_name = 'client_group' if 'client_group' in df.columns else 'Client Group'
        groups = df[col_name].unique()
        for g in groups:
            group_df = df[df[col_name] == g]
            safe_g = str(g).replace("/", "_").replace("\\", "_")
            if not safe_g.strip(): safe_g = "Ungrouped"
            
            if report_type == "All":
                excel_bytes = generate_general_excel(group_df, hide_client_group=True)
                filename = f"{safe_g}_Company_Data.xlsx"
            elif report_type == "Outstanding":
                excel_bytes = generate_beautiful_excel(group_df, hide_client_group=True)
                filename = f"{safe_g}_Outstanding_Compliance.xlsx"
            elif report_type == "Invoicing":
                excel_bytes = generate_inv_excel(group_df, hide_client_group=True)
                filename = f"{safe_g}_Invoicing_Schedule_{year}_{month_disp}.xlsx"
            
            zip_file.writestr(filename, excel_bytes)
    return zip_buffer.getvalue()

# --- 5. Dashboard ---
if choice == "📊 Dashboard":
    st.header("📊 Compliance Overview")
    df_raw = pd.read_sql("SELECT * FROM companies", engine)
    groups = pd.read_sql("SELECT group_name FROM client_groups", engine)['group_name'].tolist()
    sorted_groups = sorted([g for g in groups if isinstance(g, str)])
    
    if not df_raw.empty:
        for col in ["incorp_date", "hk_incorp_date", "nd2a_eff_date", "nd4_eff_date", "nd2a_file_date", "nd4_file_date", "nn6_eff_date", "nn6_file_date", "dissolution_date"]:
            if col in df_raw.columns: df_raw[col] = pd.to_datetime(df_raw[col], errors='coerce').dt.date
            
        today = datetime.now(HKT).date()
        outstanding_records = []
        
        for index, row in df_raw.iterrows():
            comp_rec = row.get('compliance_records')
            try: comp_rec = json.loads(comp_rec) if isinstance(comp_rec, str) else {}
            except: comp_rec = {}
            if not isinstance(comp_rec, dict): comp_rec = {}
            
            for y in active_years:
                y_str = str(y)
                y_data = comp_rec.get(y_str, {})
                df_raw.at[index, f'{y}_br_paid_by'] = y_data.get('br_paid_by', 'Firm')
                df_raw.at[index, f'{y}_br_date'] = to_date(y_data.get('br_date'))
                df_raw.at[index, f'{y}_ar_date'] = to_date(y_data.get('ar_date'))

        for _, row in df_raw.iterrows():
            name = row.get('name_en', 'Unknown')
            name_ch = str(row.get('name_ch', ''))
            group = row.get('client_group', '')
            place = str(row.get('incorp_place', ''))
            remark_val = str(row.get('remark', ''))
            if remark_val == 'None': remark_val = ""
            
            base_date = get_base_date(row)
            if not base_date: continue
            incorp_year = base_date.year
            
            for y in active_years:
                y_str = str(y)
                
                if y < incorp_year: continue 
                
                br_by = str(row.get(f'{y}_br_paid_by', 'Firm'))
                last_br = to_date(row.get(f'{y}_br_date'))
                last_ar = to_date(row.get(f'{y}_ar_date'))
                
                br_dl = get_anniv(y, base_date.month, base_date.day)
                ar_dl = br_dl + timedelta(days=42)
                
                br_status, ar_status = "🟢 Normal", "🟢 Normal"
                br_dl_str = br_dl.strftime('%Y/%m/%d')
                ar_dl_str = ar_dl.strftime('%Y/%m/%d')
                is_alert = False
                
                if br_by == 'Client':
                    if last_br:
                        br_status = "✅ Client (Recorded)"
                    else:
                        br_status = "✅ Client (Pending Record)"
                    br_dl_str = "N/A"
                elif br_by == 'N/A':
                    br_status = "✅ N/A"
                    br_dl_str = "N/A"
                else:
                    if not last_br:
                        days_diff = (br_dl - today).days
                        if days_diff < 0: br_status, is_alert = "🔴 Overdue", True
                        elif 0 <= days_diff <= 30: br_status, is_alert = "🟡 Due Soon", True
                
                if y == incorp_year:
                    ar_status = "✅ Exempt (1st Year)"
                    ar_dl_str = "N/A"
                else:
                    if not last_ar:
                        ar_days_diff = (ar_dl - today).days
                        if ar_days_diff < 0: ar_status, is_alert = "🔴 Overdue", True
                        elif 0 <= ar_days_diff <= 72: ar_status, is_alert = "🟡 Due Soon", True
                        
                if is_alert:
                    outstanding_records.append({
                        "Client Group": group,
                        "Company Name EN": name,
                        "Company Name CH": name_ch,
                        "Incorp Place": place,
                        "Year": y_str,
                        "BR Paid By": br_by,
                        "BR Status": br_status,
                        "BR Deadline": br_dl_str,
                        "AR Status": ar_status,
                        "AR Deadline": ar_dl_str,
                        "Remark": remark_val
                    })

        tab1, tab2, tab3 = st.tabs(["📊 All Companies", "🚨 Outstanding List", "🧾 Invoicing Schedule"])
        
        with tab1:
            sort_cols = [c for c in ['client_group', 'name_en', 'incorp_place'] if c in df_raw.columns]
            df_raw = df_raw.sort_values(by=sort_cols, na_position='last')
            
            t1, t2, t3, t4 = st.columns([3, 2, 2, 5])
            filter_g = t1.selectbox("🔍 Filter Group", ["All Groups"] + sorted_groups)
            if t2.button("🔄 Refresh"): st.rerun()
            df_filtered = df_raw if filter_g == "All Groups" else df_raw[df_raw['client_group'] == filter_g]
            
            if 'sel_v166' not in st.session_state: st.session_state.sel_v166 = False
            if t3.button("✅ Select All"): st.session_state.sel_v166 = True; st.rerun()
            if t4.button("🧹 Clear All"): st.session_state.sel_v166 = False; st.rerun()
            
            base_cols = [c for c in TEMPLATE_COLS if c in df_filtered.columns]
            dyn_cols = []
            for y in active_years:
                dyn_cols.extend([f"{y}_br_paid_by", f"{y}_br_date", f"{y}_ar_date"])
            
            display_cols = ["name_en", "name_ch"]
            for c in base_cols:
                if c not in display_cols and c != "remark": display_cols.append(c)
            for c in dyn_cols:
                if c not in display_cols and c != "remark": display_cols.append(c)
            display_cols.append("remark") 
                
            df_display = df_filtered[display_cols].copy()
            df_display['remark'] = df_display['remark'].fillna('')
            
            df_display.rename(columns={'name_en': 'Company Name EN', 'name_ch': 'Company Name CH'}, inplace=True)
            df_display.set_index('Company Name EN', inplace=True)
            df_display.insert(1, "Select", st.session_state.sel_v166)
            
            st.markdown(f"📈 Total: **{len(df_filtered)}** companies in current view.")
            
            col_cfg = {
                "Select": st.column_config.CheckboxColumn("Select", default=False),
                "remark": st.column_config.TextColumn("✏️ Remark")
            }
            for y in active_years:
                col_cfg[f"{y}_br_paid_by"] = st.column_config.SelectboxColumn(f"✏️ {y} BR Paid By", options=["Firm", "Client", "N/A"], required=True)
                col_cfg[f"{y}_br_date"] = st.column_config.DateColumn(f"✏️ {y} BR Date", format="YYYY/MM/DD")
                col_cfg[f"{y}_ar_date"] = st.column_config.DateColumn(f"✏️ {y} AR Date", format="YYYY/MM/DD")
            
            disabled_cols = [c for c in df_display.columns if c not in ["Select", "remark"] and not any(c.endswith(suffix) for suffix in ["_br_paid_by", "_br_date", "_ar_date"])]
            
            edit_df = st.data_editor(
                df_display, 
                column_config=col_cfg,
                disabled=disabled_cols,
                use_container_width=True, 
                key="dash_v166"
            )
            
            if st.button("💾 Save Batch Edits", key="btn_save_grid_v166"):
                try:
                    with engine.begin() as conn:
                        for comp_name, r in edit_df.iterrows():
                            row_info = df_raw[df_raw['name_en'] == comp_name].iloc[0]
                            base_dt = get_base_date(row_info)
                            inc_yr = base_dt.year if base_dt else None
                            
                            comp_dict = {}
                            for y in active_years:
                                y_str = str(y)
                                br_by = str(r.get(f'{y}_br_paid_by', 'Firm'))
                                
                                raw_br = to_date(r.get(f'{y}_br_date'))
                                raw_ar = to_date(r.get(f'{y}_ar_date'))
                                
                                if inc_yr and y < inc_yr:
                                    br_by = 'N/A'
                                    raw_br = None
                                    raw_ar = None
                                elif inc_yr and y == inc_yr:
                                    raw_ar = None
                                    
                                if br_by == 'N/A': raw_br = None
                                
                                comp_dict[y_str] = {
                                    "br_paid_by": br_by,
                                    "br_date": raw_br.strftime('%Y-%m-%d') if raw_br else None,
                                    "ar_date": raw_ar.strftime('%Y-%m-%d') if raw_ar else None
                                }
                            json_str = json.dumps(comp_dict).replace("'", "''")
                            rem_str = str(r.get('remark', '')).replace("'", "''")
                            if rem_str == 'None': rem_str = ""
                            
                            comp_name_safe = str(comp_name).replace("'", "''")
                            sql = f"UPDATE companies SET compliance_records = '{json_str}', remark = '{rem_str}' WHERE name_en = '{comp_name_safe}'"
                            conn.execute(text(sql))
                    st.success("✅ Changes saved successfully!")
                    st.rerun()
                except Exception as e:
                    st.error(f"❌ Batch save failed: {e}")
            
            selected = edit_df[edit_df["Select"] == True]
            if len(selected) > 0:
                st.info(f"✅ **{len(selected)}** companies selected for action.")
                c_act1, c_act2, c_act3, c_act4 = st.columns([2.5, 2.5, 2.5, 2.5])
                with c_act1.popover("🏢 Internal Export (All-in-One)"):
                    final_data = df_raw[df_raw['name_en'].isin(selected.index.tolist())]
                    st.download_button(label="📥 PDF (Combined)", data=generate_custom_pdf(final_data), file_name="Company_Report_Internal.pdf", mime="application/pdf", key="pdf_in_1")
                    st.download_button(label="📦 Excel (Combined)", data=generate_general_excel(final_data), file_name="Company_Data_Internal.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="exc_in_1")
                with c_act2.popover("🤝 Client Export (ZIP by Group)"):
                    st.download_button("🗂️ ZIP (PDFs by Group)", data=create_zip_pdfs(final_data, "All"), file_name="Company_Reports_ZIP.zip", mime="application/zip", key="zip_pdf_1")
                    st.download_button("🗂️ ZIP (Excels by Group)", data=create_zip_excels(final_data, "All"), file_name="Company_Data_ZIP.zip", mime="application/zip", key="zip_exc_1")
                with c_act3.popover("📄 External Export (No Group)"):
                    st.download_button("📥 PDF (No Group)", data=generate_custom_pdf(final_data, hide_client_group=True), file_name="Company_Report_External.pdf", mime="application/pdf", key="pdf_ex_1")
                    st.download_button("📦 Excel (No Group)", data=generate_general_excel(final_data, hide_client_group=True), file_name="Company_Data_External.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="exc_ex_1")
                with c_act4.popover("🧨 BATCH DELETE"):
                    st.error("🛑 DANGER ZONE"); conf_b = st.text_input("Type DELETE", key="batch_del_v166")
                    if st.button("Confirm Batch Delete", disabled=(conf_b != "DELETE"), key="btn_batch_del_v166"):
                        df_raw[~df_raw["name_en"].isin(selected.index.tolist())].to_sql('companies', engine, if_exists='replace', index=False); st.rerun()

        with tab2:
            df_alerts = pd.DataFrame(outstanding_records)
            if not df_alerts.empty:
                df_alerts = df_alerts.sort_values(by=['Client Group', 'Company Name EN', 'Incorp Place', 'Year'], na_position='last')
                
                ta1, ta2, ta3, ta4 = st.columns([3, 2, 2, 5])
                filter_alert_g = ta1.selectbox("🔍 Filter Group", ["All Groups"] + sorted_groups, key="filter_alert_g")
                if ta2.button("🔄 Refresh", key="ref_alert"): st.rerun()
                
                df_alerts_filtered = df_alerts if filter_alert_g == "All Groups" else df_alerts[df_alerts['Client Group'] == filter_alert_g]
                
                if 'sel_alert_v166' not in st.session_state: st.session_state.sel_alert_v166 = False
                if ta3.button("✅ Select All", key="sel_all_alert"): st.session_state.sel_alert_v166 = True; st.rerun()
                if ta4.button("🧹 Clear All", key="clr_all_alert"): st.session_state.sel_alert_v166 = False; st.rerun()
                
                alert_cols_order = ["Company Name EN", "Company Name CH", "Client Group", "Incorp Place", "Year", "BR Paid By", "BR Status", "BR Deadline", "AR Status", "AR Deadline", "Remark"]
                df_alerts_display = df_alerts_filtered[alert_cols_order].copy()
                df_alerts_display.set_index('Company Name EN', inplace=True)
                df_alerts_display.insert(1, "Select", st.session_state.sel_alert_v166)
                
                st.markdown(f"📈 Total: **{len(df_alerts_display)}** tasks in current view.")
                
                alert_edit = st.data_editor(
                    df_alerts_display,
                    column_config={"Select": st.column_config.CheckboxColumn("Select", default=False)}, 
                    use_container_width=True,
                    disabled=[c for c in df_alerts_display.columns if c != "Select"],
                    key="alert_grid_v166"
                )
                
                selected_alerts = df_alerts_display[alert_edit["Select"] == True]
                export_target = selected_alerts.reset_index() if len(selected_alerts) > 0 else None
                
                if export_target is not None:
                    st.info(f"✅ **{len(export_target)}** tasks selected for export.")
                    ca1, ca2, ca3 = st.columns([3, 3, 4])
                    with ca1.popover("🏢 Internal Export (All-in-One)"):
                        st.download_button("📥 PDF (Combined)", data=generate_outstanding_pdf(export_target), file_name="Outstanding_Internal.pdf", mime="application/pdf", key="pdf_in_2")
                        st.download_button("📦 Excel (Combined)", data=generate_beautiful_excel(export_target), file_name="Outstanding_Internal.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="exc_in_2")
                    with ca2.popover("🤝 Client Export (ZIP by Group)"):
                        st.download_button("🗂️ ZIP (PDFs by Group)", data=create_zip_pdfs(export_target, "Outstanding"), file_name="Outstanding_ZIP.zip", mime="application/zip", key="zip_pdf_2")
                        st.download_button("🗂️ ZIP (Excels by Group)", data=create_zip_excels(export_target, "Outstanding"), file_name="Outstanding_ZIP.zip", mime="application/zip", key="zip_exc_2")
                    with ca3.popover("📄 External Export (No Group)"):
                        st.download_button("📥 PDF (No Group)", data=generate_outstanding_pdf(export_target, hide_client_group=True), file_name="Outstanding_External.pdf", mime="application/pdf", key="pdf_ex_2")
                        st.download_button("📦 Excel (No Group)", data=generate_beautiful_excel(export_target, hide_client_group=True), file_name="Outstanding_External.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="exc_ex_2")
                else:
                    st.info("💡 Select checkboxes to export specific companies.")
            else:
                st.success("🎉 No outstanding tasks at the moment!")
                
        with tab3:
            ti1, ti2, ti3, ti4, ti5, ti6, ti7 = st.columns([2, 1.5, 1.5, 1.5, 1, 1, 3])
            filter_inv_g = ti1.selectbox("🔍 Filter Group", ["All Groups"] + sorted_groups, key="filter_inv_g")
            target_year = ti2.selectbox("📅 Target Year", active_years, index=active_years.index(current_system_year), key="inv_year")
            
            start_month = ti3.selectbox("📆 From Month", range(1, 13), index=0, key="inv_start_m")
            end_month = ti4.selectbox("📆 To Month", range(1, 13), index=11, key="inv_end_m")
            month_str = f"{start_month}-{end_month}" if start_month != end_month else str(start_month)
            
            if ti5.button("🔄 Refresh", key="ref_inv"): st.rerun()
            
            if 'sel_inv_v166' not in st.session_state: st.session_state.sel_inv_v166 = False
            if ti6.button("✅ Select All", key="sel_all_inv"): st.session_state.sel_inv_v166 = True; st.rerun()
            if ti7.button("🧹 Clear All", key="clr_all_inv"): st.session_state.sel_inv_v166 = False; st.rerun()
            
            inv_records = []
            for index, row in df_raw.iterrows():
                place = str(row.get('incorp_place', ''))
                is_hk_reg = str(row.get('is_hk_registered', 'False')).strip().lower() in ['true', 'yes', 'y', '1']
                
                if place != 'HK' and not is_hk_reg: continue
                
                base_date = get_base_date(row)
                if not base_date: continue
                incorp_year = base_date.year
                
                if target_year < incorp_year: continue
                
                b_month = base_date.month
                if start_month <= end_month:
                    if not (start_month <= b_month <= end_month): continue
                else: 
                    if not (b_month >= start_month or b_month <= end_month): continue
                
                group = row.get('client_group', '')
                if filter_inv_g != "All Groups" and group != filter_inv_g: continue
                
                name = row.get('name_en', 'Unknown')
                name_ch = str(row.get('name_ch', ''))
                remark_val = str(row.get('remark', ''))
                if remark_val == 'None': remark_val = ""
                
                comp_rec_str = str(row.get('compliance_records', '{}'))
                try: rec_dict = json.loads(comp_rec_str) if isinstance(comp_rec_str, str) else {}
                except: rec_dict = {}
                y_data = rec_dict.get(str(target_year), {})
                
                br_by = str(y_data.get('br_paid_by', 'Firm'))
                
                br_dl = get_anniv(target_year, base_date.month, base_date.day)
                ar_dl = br_dl + timedelta(days=42)
                
                br_dl_str = br_dl.strftime('%Y/%m/%d')
                if target_year == incorp_year:
                    ar_dl_str = "Exempt (1st Year)"
                else:
                    ar_dl_str = ar_dl.strftime('%Y/%m/%d')
                    
                inv_records.append({
                    "Company Name EN": name,
                    "Company Name CH": name_ch,
                    "Client Group": group,
                    "Incorp Place": place,
                    "Year": str(target_year),
                    "Anniversary (MM/DD)": base_date.strftime('%m/%d'),
                    "BR Paid By": br_by,
                    "BR Deadline": br_dl_str,
                    "AR Deadline": ar_dl_str,
                    "Remark": remark_val
                })
                
            df_inv = pd.DataFrame(inv_records)
            if not df_inv.empty:
                df_inv = df_inv.sort_values(by=['Client Group', 'Company Name EN', 'Incorp Place', 'Year', 'Anniversary (MM/DD)'])
                
                inv_cols_order = ["Company Name EN", "Company Name CH", "Client Group", "Incorp Place", "Year", "Anniversary (MM/DD)", "BR Paid By", "BR Deadline", "AR Deadline", "Remark"]
                df_inv_display = df_inv[inv_cols_order].copy()
                df_inv_display.set_index('Company Name EN', inplace=True)
                df_inv_display.insert(1, "Select", st.session_state.sel_inv_v166)
                
                st.markdown(f"📈 Total: **{len(df_inv_display)}** companies for Invoicing in current view.")
                
                inv_edit = st.data_editor(
                    df_inv_display,
                    column_config={"Select": st.column_config.CheckboxColumn("Select", default=False)}, 
                    use_container_width=True,
                    disabled=[c for c in df_inv_display.columns if c != "Select"],
                    key="inv_grid_v166"
                )
                
                selected_inv = df_inv_display[inv_edit["Select"] == True]
                export_target_inv = selected_inv.reset_index() if len(selected_inv) > 0 else None
                
                if export_target_inv is not None:
                    st.info(f"✅ **{len(export_target_inv)}** companies selected for export.")
                    ci1, ci2, ci3 = st.columns([3, 3, 4])
                    with ci1.popover("🏢 Internal Export (All-in-One)"):
                        st.download_button("📥 PDF (Combined)", data=generate_inv_pdf(export_target_inv, str(target_year), month_str), file_name=f"Invoicing_Internal_{target_year}_{month_str}.pdf", mime="application/pdf", key="pdf_in_3")
                        st.download_button("📦 Excel (Combined)", data=generate_inv_excel(export_target_inv), file_name=f"Invoicing_Internal_{target_year}_{month_str}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="exc_in_3")
                    with ci2.popover("🤝 Client Export (ZIP by Group)"):
                        st.download_button("🗂️ ZIP (PDFs by Group)", data=create_zip_pdfs(export_target_inv, "Invoicing", str(target_year), month_str), file_name=f"Invoicing_ZIP_{target_year}_{month_str}.zip", mime="application/zip", key="zip_pdf_3")
                        st.download_button("🗂️ ZIP (Excels by Group)", data=create_zip_excels(export_target_inv, "Invoicing", str(target_year), month_str), file_name=f"Invoicing_ZIP_{target_year}_{month_str}.zip", mime="application/zip", key="zip_exc_3")
                    with ci3.popover("📄 External Export (No Group)"):
                        st.download_button("📥 PDF (No Group)", data=generate_inv_pdf(export_target_inv, str(target_year), month_str, hide_client_group=True), file_name=f"Invoicing_External_{target_year}_{month_str}.pdf", mime="application/pdf", key="pdf_ex_3")
                        st.download_button("📦 Excel (No Group)", data=generate_inv_excel(export_target_inv, hide_client_group=True), file_name=f"Invoicing_External_{target_year}_{month_str}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="exc_ex_3")
                else:
                    st.info("💡 Select checkboxes to export specific companies.")
            else:
                st.success("🎉 No companies found for this invoicing period.")
                
    else: st.info("No records found.")

# --- 6. Company Register ---
elif choice == "🏢 Company Register":
    st.title("🏢 Company Records Management")
    mode = st.radio("Mode", ["🆕 Add New", "✏️ Edit Existing", "📋 Copy Existing"], horizontal=True)
    df_all = pd.read_sql("SELECT * FROM companies", engine)
    groups = pd.read_sql("SELECT group_name FROM client_groups", engine)['group_name'].tolist()
    sorted_groups = sorted([g for g in groups if isinstance(g, str)])
    MIN_DATE = datetime(1900, 1, 1)

    d = {'cg': "", 'en': "", 'ch': "", 'place': "", 'p_oth': "", 'idate': None, 'ci': "", 'is_hk_reg': False, 'hk_idate': None, 'hk_ci': "", 'br': "", 'type': "", 'ra': "", 'ca': "", 'rl': "", 'sl': "", 'cl': "", 'n2e': None, 'n2f': None, 'n2d': False, 'n4e': None, 'n4f': None, 'n4d': False, 'nn6_e': None, 'nn6_f': None, 'nn6_d': False, 'dis': None, 'rem': "", 'comp_rec_dict': {}}
    target_name = None
    if mode in ["✏️ Edit Existing", "📋 Copy Existing"] and not df_all.empty:
        df_all = df_all.sort_values(by=['name_en', 'incorp_place'], na_position='last')
        sorted_companies = df_all['name_en'].tolist()
        
        target_name = st.selectbox("Select Company", [""] + sorted_companies)
        if target_name != "":
            row = df_all[df_all['name_en'] == target_name].iloc[0]
            comp_rec = row.get('compliance_records')
            try: comp_rec = json.loads(comp_rec) if isinstance(comp_rec, str) else {}
            except: comp_rec = {}
            if not isinstance(comp_rec, dict): comp_rec = {}
            
            rem_val = str(row.get('remark', ''))
            if rem_val == 'None': rem_val = ""
            
            d = {'cg': row.get('client_group', ""), 'en': row.get('name_en', ""), 'ch': row.get('name_ch', ""), 'place': row.get('incorp_place', ""), 'p_oth': row.get('incorp_place_others', ""), 'idate': row.get('incorp_date'), 'ci': row.get('ci_no', ""), 'is_hk_reg': str(row.get('is_hk_registered', "")) == 'True', 'hk_idate': row.get('hk_incorp_date'), 'hk_ci': row.get('hk_ci_no', ""), 'br': row.get('br_no', ""), 'type': row.get('co_type', ""), 'ra': row.get('reg_addr', ""), 'ca': row.get('corres_addr', ""), 'rl': row.get('round_loc', ""), 'sl': row.get('sign_loc', ""), 'cl': row.get('seal_loc', ""), 'n2e': row.get('nd2a_eff_date'), 'n2f': row.get('nd2a_file_date'), 'n2d': str(row.get('nd2a_download', "")) == 'True', 'n4e': row.get('nd4_eff_date'), 'n4f': row.get('nd4_file_date'), 'n4d': str(row.get('nd4_download', "")) == 'True', 'nn6_e': row.get('nn6_eff_date'), 'nn6_f': row.get('nn6_file_date'), 'nn6_d': str(row.get('nn6_download', "")) == 'True', 'dis': row.get('dissolution_date'), 'rem': rem_val, 'comp_rec_dict': comp_rec}
            if mode == "📋 Copy Existing": d['en'], d['ch'] = "", ""

    st.header("General Information")
    c1, c2 = st.columns(2)
    with c1: st.markdown("⚠️ Company English Name :red[(Required!)]"); name_en = st.text_input("EN", value=d['en'], label_visibility="collapsed")
    with c2: st.markdown("Company Chinese Name"); name_ch = st.text_input("CH", value=d['ch'], label_visibility="collapsed")
    st.markdown("⚠️ Select Client Group :red[(Required!)]")
    client_group = st.selectbox("Group", [""] + sorted_groups, index=(sorted_groups.index(d['cg'])+1 if d['cg'] in sorted_groups else 0), label_visibility="collapsed")
    st.write("---") 
    
    place_options = ["", "HK", "BVI", "Cayman Island", "Others"]
    st.markdown("⚠️ Place of Incorporation :red[(Required!)]")
    inc_place = st.selectbox("Place", place_options, index=(place_options.index(d['place']) if d['place'] in place_options else 0), label_visibility="collapsed")
    
    place_others = ""
    if inc_place == "Others": 
        st.markdown("⚠️ Specify Others :red[(Required!)]"); place_others = st.text_input("Others_Input", value=d['p_oth'], label_visibility="collapsed")
    
    if inc_place:
        disp_place = "Others" if inc_place == "Others" else inc_place
        c3, c4 = st.columns(2)
        with c3: st.markdown(f"⚠️ {disp_place} Incorp. Date (YYYY/MM/DD) :red[(Required!)]"); inc_date = st.date_input("Date", value=to_date(d['idate']), min_value=MIN_DATE, label_visibility="collapsed", format="YYYY/MM/DD")
        with c4: st.markdown(f"⚠️ {disp_place} CI Number :red[(Required!)]"); ci_no = st.text_input("CI", value=d['ci'], label_visibility="collapsed")
    else:
        inc_date = None
        ci_no = ""
    
    is_hk_reg, hk_idate, hk_ci, br_no = False, None, "", ""
    if inc_place == "HK":
        st.markdown("⚠️ HK BR Number :red[(Required!)]")
        br_no = st.text_input("BR", value=d['br'], label_visibility="collapsed")
    elif inc_place in ["BVI", "Cayman Island", "Others"]:
        st.write("---")
        is_hk_reg = st.checkbox("Registered as Non-Hong Kong Company in HK?", value=d['is_hk_reg'])
        if is_hk_reg:
            st.info("📌 Hong Kong Registration Details")
            hk1, hk2 = st.columns(2)
            with hk1: st.markdown("⚠️ HK Incorp. Date (YYYY/MM/DD) :red[(Required!)]"); hk_idate = st.date_input("HK_Date", value=to_date(d['hk_idate']), min_value=MIN_DATE, label_visibility="collapsed", format="YYYY/MM/DD")
            with hk2: st.markdown("⚠️ HK CI Number :red[(Required!)]"); hk_ci = st.text_input("HK_CI", value=d['hk_ci'], label_visibility="collapsed")
            st.markdown("⚠️ HK BR Number :red[(Required!)]")
            br_no = st.text_input("BR", value=d['br'], label_visibility="collapsed")

    st.write("---") 
    type_options = ["", "Private Company", "Public Company", "Guarantee", "Individual Business", "Non-Hong Kong Company"]
    st.markdown("⚠️ Company Type :red[(Required!)]"); co_type = st.selectbox("Type", type_options, index=(type_options.index(d['type']) if d['type'] in type_options else 0), label_visibility="collapsed")

    # ==================== 📅 Dynamic Annual Obligations ====================
    updated_comp_json = {}
    if inc_place == "HK" or is_hk_reg:
        st.write("---"); st.header("📅 Annual Obligations (2025 - Present)")
        base = hk_idate if is_hk_reg else inc_date
        comp_json_load = d.get('comp_rec_dict', {})
        incorp_year = base.year if base else None
        
        if base:
            today_cal = datetime.now(HKT).date()
            for y in active_years:
                with st.expander(f"📌 {y} 年度申報 (Year {y})", expanded=True):
                    y_str = str(y)
                    y_data = comp_json_load.get(y_str, {})
                    
                    if incorp_year and y < incorp_year:
                        st.info(f"### ⚪ Year {y}: Not Incorporated Yet (此年份公司尚未成立)")
                    elif incorp_year and y == incorp_year:
                        st.success(f"### ✅ AR Deadline ({y}): Exempt (首年豁免周年申報)")
                        
                    nxt_br = get_anniv(y, base.month, base.day)
                    nxt_ar = nxt_br + timedelta(days=42)
                    br_days = (nxt_br - today_cal).days
                    ar_days = (nxt_ar - today_cal).days
                    
                    col_m1, col_m2 = st.columns(2)
                    with col_m1:
                        if y_data.get('br_paid_by') == "Client":
                            if to_date(y_data.get('br_date')):
                                st.success(f"### 🟢 BR Fee Deadline ({y}): **{nxt_br.strftime('%Y/%m/%d')}**\n\n✅ **Client (Recorded / 已存檔)**")
                            else:
                                st.success(f"### 🟢 BR Fee Deadline ({y}): **{nxt_br.strftime('%Y/%m/%d')}**\n\n✅ **Client (Pending Record / 待補紀錄)**")
                        elif y_data.get('br_paid_by') == "N/A": 
                            st.success(f"### 🟢 BR Fee Deadline ({y}): **{nxt_br.strftime('%Y/%m/%d')}**\n\n✅ **N/A**")
                        else:
                            if y > today_cal.year: st.info(f"### 🔵 BR Fee Deadline ({y}): **{nxt_br.strftime('%Y/%m/%d')}**\n\n🕒 Not yet due")
                            elif br_days < 0: st.error(f"### 🚨 BR Fee Deadline ({y}): **{nxt_br.strftime('%Y/%m/%d')}**\n\n⚠️ **Overdue by {abs(br_days)} days**")
                            elif br_days <= 30: st.warning(f"### ⏳ BR Fee Deadline ({y}): **{nxt_br.strftime('%Y/%m/%d')}**\n\n⏰ **Due in {br_days} days**")
                            else: st.success(f"### 🟢 BR Fee Deadline ({y}): **{nxt_br.strftime('%Y/%m/%d')}**\n\n✅ **Status: Normal**")
                                
                    with col_m2:
                        if y == incorp_year:
                            st.success(f"### ✅ AR Deadline ({y}): Exempt")
                        elif y > today_cal.year: st.info(f"### 🔵 AR Deadline ({y}): **{nxt_ar.strftime('%Y/%m/%d')}**\n\n🕒 Not yet due")
                        elif ar_days < 0: st.error(f"### 🚨 AR Deadline ({y}): **{nxt_ar.strftime('%Y/%m/%d')}**\n\n⚠️ **Overdue by {abs(ar_days)} days**")
                        elif ar_days <= 72: st.warning(f"### ⏳ AR Deadline ({y}): **{nxt_ar.strftime('%Y/%m/%d')}**\n\n⏰ **Due in {ar_days} days**")
                        else: st.success(f"### 🟢 AR Deadline ({y}): **{nxt_ar.strftime('%Y/%m/%d')}**\n\n✅ **Status: Normal**")
                
                    o1, o2, o3 = st.columns([3, 4, 4])
                    with o1:
                        opt_list = ["Firm", "Client", "N/A"]
                        br_paid_by_idx = opt_list.index(y_data.get('br_paid_by', 'Firm')) if y_data.get('br_paid_by', 'Firm') in opt_list else 0
                        br_by = st.selectbox(f"BR Paid By ({y})", opt_list, index=br_paid_by_idx, key=f"br_by_{y}")
                    with o2: 
                        if br_by == "N/A":
                            st.text_input(f"BR Paid Date ({y}) (YYYY/MM/DD)", value="N/A", disabled=True, key=f"br_dt_dis_{y}")
                            l_br = None
                        else:
                            l_br = st.date_input(f"BR Paid Date ({y})", value=to_date(y_data.get('br_date')), min_value=MIN_DATE, key=f"br_dt_{y}", format="YYYY/MM/DD")
                    with o3: 
                        if y == incorp_year:
                            st.text_input(f"AR Filed Date ({y}) (YYYY/MM/DD)", value="N/A (Exempt)", disabled=True, key=f"ar_dt_dis_{y}")
                            l_ar = None
                        else:
                            l_ar = st.date_input(f"AR Filed Date ({y})", value=to_date(y_data.get('ar_date')), min_value=MIN_DATE, key=f"ar_dt_{y}", format="YYYY/MM/DD")
                        
                    updated_comp_json[y_str] = {
                        "br_paid_by": br_by,
                        "br_date": l_br.strftime('%Y-%m-%d') if l_br else None,
                        "ar_date": l_ar.strftime('%Y-%m-%d') if l_ar else None
                    }

    n2e, n2f, n2d = d['n2e'], d['n2f'], d['n2d']
    n4e, n4f, n4d = d['n4e'], d['n4f'], d['n4d']
    nn6_e, nn6_f, nn6_d = d['nn6_e'], d['nn6_f'], d['nn6_d']

    if inc_place == "HK":
        st.write("---"); st.header("📝 Compliance Filings (Local Company)")
        st.subheader("📑 Company Secretary Appointment (ND2A)")
        cc1, cc2, cc3, cc4 = st.columns([3, 3, 3, 1])
        with cc1: n2e = st.date_input("Effective Date (Appt)", value=to_date(d['n2e']), min_value=MIN_DATE, key="n2e_v166", format="YYYY/MM/DD")
        with cc2: n2f = st.date_input("Filing Date (ND2A)", value=to_date(d['n2f']), min_value=MIN_DATE, key="n2f_v166", format="YYYY/MM/DD")
        with cc3:
            st.info("Statutory Period: 15 days")
            if n2e: n2_deadline = (n2e + timedelta(days=15)); st.markdown(f"**Deadline: :red[{n2_deadline.strftime('%Y/%m/%d')}]**") 
        with cc4: n2d = st.checkbox("Downloaded", value=d['n2d'], key="n2d_v166")
        
        st.subheader("📑 Company Secretary Resignation (ND4)")
        cc5, cc6, cc7, cc8 = st.columns([3, 3, 3, 1])
        with cc5: n4e = st.date_input("Effective Date (Resign)", value=to_date(d['n4e']), min_value=MIN_DATE, key="n4e_v166", format="YYYY/MM/DD")
        with cc6: n4f = st.date_input("Filing Date (ND4)", value=to_date(d['n4f']), min_value=MIN_DATE, key="n4f_v166", format="YYYY/MM/DD")
        with cc7:
            st.info("Statutory Period: 15 days")
            if n4e: n4_deadline = (n4e + timedelta(days=15)); st.markdown(f"**Deadline: :red[{n4_deadline.strftime('%Y/%m/%d')}]**") 
        with cc8: n4d = st.checkbox("Downloaded", value=d['n4d'], key="n4d_v166")
        
    elif is_hk_reg:
        st.write("---"); st.header("📝 Compliance Filings (Non-HK Company)")
        st.subheader("📑 Secretary & Director Changes (NN6)")
        c_nn1, c_nn2, c_nn3, c_nn4 = st.columns([3, 3, 3, 1])
        with c_nn1: nn6_e = st.date_input("Effective Date", value=to_date(d['nn6_e']), min_value=MIN_DATE, key="nn6_e_v166", format="YYYY/MM/DD")
        with c_nn2: nn6_f = st.date_input("Filing Date (NN6)", value=to_date(d['nn6_f']), min_value=MIN_DATE, key="nn6_f_v166", format="YYYY/MM/DD")
        with c_nn3:
            st.info("Statutory Period: 1 Month")
            if nn6_e:
                nn6_deadline = add_one_month(nn6_e)
                st.markdown(f"**Deadline: :red[{nn6_deadline.strftime('%Y/%m/%d')}]**")
        with c_nn4: nn6_d = st.checkbox("Downloaded", value=d['nn6_d'], key="nn6_d_v166")

    st.write("---"); st.subheader("📍 Address & Contact")
    ca1, ca2 = st.columns(2)
    with ca1: st.markdown("⚠️ Registered Office Address :red[(Required!)]"); reg_addr = st.text_area("Reg", value=d['ra'], label_visibility="collapsed")
    with ca2: st.markdown("⚠️ Correspondence Address :red[(Required!)]"); corres_addr = st.text_area("Corres", value=d['ca'], label_visibility="collapsed")
    st.subheader("📔 Seal Storage")
    l1, l2, l3 = st.columns(3)
    with l1: st.markdown("⚠️ Round Chop Location :red[(Required!)]"); round_l = st.text_input("Round", value=d['rl'], label_visibility="collapsed")
    with l2: st.markdown("⚠️ Signature Chop Location :red[(Required!)]"); sign_l = st.text_input("Sign", value=d['sl'], label_visibility="collapsed")
    with l3: st.markdown("⚠️ Common Seal Location :red[(Required!)]"); common_l = st.text_input("Seal", value=d['cl'], label_visibility="collapsed")
    st.write("---"); st.markdown("Company Dissolution Date"); dis_date = st.date_input("Dissolution", value=to_date(d['dis']), min_value=MIN_DATE, label_visibility="collapsed", format="YYYY/MM/DD")
    
    st.write("---"); st.subheader("📌 Remarks")
    remark_input = st.text_area("Remark / 備註", value=d['rem'], help="此備註會同步顯示於報告及總覽表格中。")
    
    row_v166 = {'client_group': client_group, 'name_en': name_en, 'name_ch': name_ch, 'incorp_place': inc_place, 'incorp_place_others': place_others, 'incorp_date': inc_date, 'ci_no': ci_no, 'is_hk_registered': is_hk_reg, 'hk_incorp_date': hk_idate, 'hk_ci_no': hk_ci, 'br_no': br_no, 'co_type': co_type, 'reg_addr': reg_addr, 'corres_addr': corres_addr, 'round_loc': round_l, 'sign_loc': sign_l, 'seal_loc': common_l, 'nd2a_eff_date': n2e, 'nd2a_file_date': n2f, 'nd2a_download': n2d, 'nd4_eff_date': n4e, 'nd4_file_date': n4f, 'nd4_download': n4d, 'nn6_eff_date': nn6_e, 'nn6_file_date': nn6_f, 'nn6_download': nn6_d, 'dissolution_date': dis_date, 'remark': remark_input, 'compliance_records': json.dumps(updated_comp_json)}
    
    mandatory_fields = {"Client Group": client_group, "English Name": name_en, "Place": inc_place, "Company Type": co_type, "Registered Address": reg_addr, "Correspondence Address": corres_addr, "Round Chop Location": round_l, "Signature Chop Location": sign_l, "Common Seal Location": common_l}
    
    if inc_place:
        mandatory_fields[f"{inc_place} Incorp Date"] = inc_date
        mandatory_fields[f"{inc_place} CI Number"] = ci_no
        if inc_place == "Others": mandatory_fields["Specify Others"] = place_others
        if inc_place == "HK": mandatory_fields["BR Number"] = br_no
        
    if is_hk_reg:
        mandatory_fields["HK Incorp Date"] = hk_idate
        mandatory_fields["HK CI Number"] = hk_ci
        mandatory_fields["HK BR Number"] = br_no
        
    missing = [k for k, v in mandatory_fields.items() if not v or str(v).strip() == ""]

    if mode in ["🆕 Add New", "📋 Copy Existing"]:
        if st.button("💾 Save To Cloud", key="btn_save_v166"):
            if missing: st.error(f"❌ Missing mandatory fields: {', '.join(missing)}")
            else:
                try:
                    pd.DataFrame([row_v166]).to_sql('companies', engine, if_exists='append', index=False)
                    st.success("✅ Success!"); st.rerun()
                except Exception as save_err:
                    st.error(f"❌ Save Failed! Error details: {save_err}")
    else:
        u_col, d_col = st.columns(2)
        with u_col.popover("🆙 Update"):
            if st.button("Confirm Update", key="btn_update_v166"):
                if missing: st.error(f"❌ Missing mandatory fields: {', '.join(missing)}")
                else:
                    try:
                        df_backup = df_all.copy() 
                        df_all[df_all['name_en'] != target_name].to_sql('companies', engine, if_exists='replace', index=False)
                        pd.DataFrame([row_v166]).to_sql('companies', engine, if_exists='append', index=False)
                        st.success("✅ Updated!"); st.rerun()
                    except Exception as trans_err:
                        df_backup.to_sql('companies', engine, if_exists='replace', index=False)
                        st.error(f"🛑 SQL Error Detected! Rollback completed. Details: {trans_err}")
        with d_col.popover("🚨 DELETE"):
            st.error(f"Delete {target_name}?"); conf_s = st.text_input("Type DELETE", key="single_del_v166")
            if st.button("Confirm Delete Company", disabled=(conf_s != "DELETE"), key="btn_del_single_v166"):
                df_all[df_all['name_en'] != target_name].to_sql('companies', engine, if_exists='replace', index=False); st.rerun()

# --- 7. Group Management ---
elif choice == "⚙️ Group Management":
    st.header("⚙️ Group Management")
    new_g = st.text_input("New Group Name", key="new_group_input_v166")
    if st.button("Add Group", key="btn_add_group_v166"): pd.DataFrame([{'group_name': new_g}]).to_sql('client_groups', engine, if_exists='append', index=False); st.rerun()
    st.write("---")
    g_df = pd.read_sql("SELECT * FROM client_groups", engine)
    if not g_df.empty:
        g_df = g_df.sort_values(by=['group_name'], na_position='last')
        target = st.selectbox("Select Group", g_df['group_name'].tolist(), key="select_group_manage_v166")
        c1, c2 = st.columns(2)
        with c1.popover("✏️ Rename Group"):
            ren = st.text_input("New Name:", key="rename_input_v166")
            conf_r = st.text_input("Type RENAME", key="rename_confirm_text_v166")
            if st.button("Confirm Rename", disabled=(conf_r != "RENAME"), key="btn_group_rename_v166"):
                comp_df = pd.read_sql("SELECT * FROM companies", engine)
                comp_df.loc[comp_df['client_group'] == target, 'client_group'] = ren
                comp_df.to_sql('companies', engine, if_exists='replace', index=False)
                g_df.replace({target: ren}).to_sql('client_groups', engine, if_exists='replace', index=False); st.rerun()
        with c2.popover("🗑️ Delete Group"):
            if st.button("Confirm Delete Group", key="btn_group_delete_v166"): 
                g_df[g_df['group_name'] != target].to_sql('client_groups', engine, if_exists='replace', index=False); st.rerun()

# --- 8. Data Exchange ---
elif choice == "📤 Data Exchange":
    st.header("📤 Data Exchange")
    c1, c2 = st.columns(2)
    
    df_template = pd.DataFrame(columns=TEMPLATE_COLS)
    for y in active_years:
        df_template[f"{y} BR Paid By"] = ""
        df_template[f"{y} BR Date"] = ""
        df_template[f"{y} AR Date"] = ""
    base_cols_t = [c for c in TEMPLATE_COLS if c != 'remark']
    
    dyn_cols_t = []
    for y in active_years:
        dyn_cols_t.extend([f"{y} BR Paid By", f"{y} BR Date", f"{y} AR Date"])
        
    df_template = df_template[base_cols_t + dyn_cols_t + ['remark']]
    
    buf_t = io.BytesIO(); df_template.to_excel(buf_t, index=False); c1.download_button(label="📥 Template", data=buf_t.getvalue(), file_name="Template.xlsx")
    
    df_db = pd.read_sql("SELECT * FROM companies", engine); df_export = df_db.copy()
    sort_cols = [c for c in ['client_group', 'name_en', 'incorp_place'] if c in df_export.columns]
    df_export = df_export.sort_values(by=sort_cols, na_position='last')
    
    for y in active_years:
        df_export[f"{y} BR Paid By"] = "Firm"
        df_export[f"{y} BR Date"] = ""
        df_export[f"{y} AR Date"] = ""
        
    for idx, row in df_export.iterrows():
        comp_rec_str = str(row.get('compliance_records', '{}'))
        try: rec_dict = json.loads(comp_rec_str)
        except: rec_dict = {}
        if isinstance(rec_dict, dict):
            for y in active_years:
                y_str = str(y)
                y_data = rec_dict.get(y_str, {})
                df_export.at[idx, f'{y} BR Paid By'] = y_data.get('br_paid_by', 'Firm')
                
                br_d = y_data.get('br_date', '')
                ar_d = y_data.get('ar_date', '')
                if br_d: br_d = br_d.replace('-', '/')
                if ar_d: ar_d = ar_d.replace('-', '/')
                
                df_export.at[idx, f'{y} BR Date'] = br_d
                df_export.at[idx, f'{y} AR Date'] = ar_d
                
    base_cols = [c for c in TEMPLATE_COLS if c in df_export.columns and c != 'remark']
    
    dyn_cols = []
    for y in active_years:
        dyn_cols.extend([f"{y} BR Paid By", f"{y} BR Date", f"{y} AR Date"])
    
    df_export = df_export[base_cols + dyn_cols + ['remark']]
    for col in ["incorp_date", "hk_incorp_date", "nd2a_eff_date", "nd2a_file_date", "nd4_eff_date", "nd4_file_date", "nn6_eff_date", "nn6_file_date", "dissolution_date"]:
        if col in df_export.columns: df_export[col] = pd.to_datetime(df_export[col], errors='coerce').dt.strftime('%Y/%m/%d')
    buf_e = io.BytesIO(); df_export.to_excel(buf_e, index=False)
    c2.download_button(label="📦 Export All", data=buf_e.getvalue(), file_name="Backup.xlsx", key="btn_export_all_v166")
    st.write("---")
    
    up = st.file_uploader("Upload XLSX to Review Changes", type=["xlsx"], key="file_uploader_v166")
    if up:
        try:
            up_df = pd.read_excel(up, engine='openpyxl', keep_default_na=False)
            existing_df = pd.read_sql("SELECT * FROM companies", engine)
            
            for col in ["incorp_date", "hk_incorp_date", "nd2a_eff_date", "nd2a_file_date", "nd4_eff_date", "nd4_file_date", "nn6_eff_date", "nn6_file_date", "dissolution_date"]:
                if col in up_df.columns: up_df[col] = pd.to_datetime(up_df[col], errors='coerce').dt.date
                if col in existing_df.columns: existing_df[col] = pd.to_datetime(existing_df[col], errors='coerce').dt.date

            validation_errors = []
            for idx, row_new in up_df.iterrows():
                excel_row = idx + 2 
                name_en = str(row_new.get('name_en', 'Unknown')).strip()
                place = str(row_new.get('incorp_place', '')).strip()
                is_hk_reg = str(row_new.get('is_hk_registered', 'False')).strip().lower() in ['true', 'yes', 'y', '1']
                
                missing_fields = []
                if not str(row_new.get('client_group', '')).strip(): missing_fields.append("Client Group")
                if not name_en: missing_fields.append("English Name")
                if not place: missing_fields.append("Place of Incorporation")
                if not str(row_new.get('co_type', '')).strip(): missing_fields.append("Company Type")
                if not str(row_new.get('reg_addr', '')).strip(): missing_fields.append("Registered Address")
                if not str(row_new.get('corres_addr', '')).strip(): missing_fields.append("Correspondence Address")
                if not str(row_new.get('round_loc', '')).strip(): missing_fields.append("Round Chop Location")
                if not str(row_new.get('sign_loc', '')).strip(): missing_fields.append("Signature Chop Location")
                if not str(row_new.get('seal_loc', '')).strip(): missing_fields.append("Common Seal Location")
                
                if place:
                    if pd.isnull(row_new.get('incorp_date')): missing_fields.append(f"{place} Incorp Date")
                    if not str(row_new.get('ci_no', '')).strip(): missing_fields.append(f"{place} CI Number")
                    if place == 'Others' and not str(row_new.get('incorp_place_others', '')).strip(): missing_fields.append("Specify Others")
                    
                if is_hk_reg:
                    if pd.isnull(row_new.get('hk_incorp_date')): missing_fields.append("HK Incorp Date")
                    if not str(row_new.get('hk_ci_no', '')).strip(): missing_fields.append("HK CI Number")
                
                if missing_fields:
                    validation_errors.append(f"**Row {excel_row} ({name_en})** - Missing mandatory fields: :red[{', '.join(missing_fields)}]")

            if validation_errors:
                st.error("🛑 **Upload Failed: Mandatory fields are missing in the Excel file. Please correct them and re-upload.**")
                for err in validation_errors:
                    st.markdown(f"- {err}")
            else:
                def get_anchor(r):
                    place = str(r.get('incorp_place', '')).strip()
                    name = str(r.get('name_en', '')).strip()
                    return f"NAME_{name}_PLACE_{place}"

                up_df['_anchor'] = up_df.apply(get_anchor, axis=1)
                existing_df['_anchor'] = existing_df.apply(get_anchor, axis=1)
                
                diff_list = []
                for _, row_new in up_df.iterrows():
                    anchor_val = row_new['_anchor']
                    en_name = row_new.get('name_en', 'Unknown')
                    old_row = existing_df[existing_df['_anchor'] == anchor_val]
                    
                    if not old_row.empty:
                        old_row = old_row.iloc[0]
                        for col in TEMPLATE_COLS:
                            if col == 'compliance_records': continue
                            old_v = clean_val(old_row.get(col, ""))
                            new_v = clean_val(row_new.get(col, ""))
                            if old_v != new_v:
                                diff_list.append({"Company": en_name, "Field": col, "Old Value": old_v if old_v else "N/A", "New Value": new_v if new_v else "N/A"})
                                
                        comp_rec_str = str(old_row.get('compliance_records', '{}'))
                        try: rec_dict = json.loads(comp_rec_str)
                        except: rec_dict = {}
                        
                        for y in active_years:
                            y_str = str(y)
                            y_data = rec_dict.get(y_str, {})
                            
                            old_br_by = str(y_data.get('br_paid_by', 'Firm'))
                            old_br_dt = str(y_data.get('br_date', ''))
                            old_ar_dt = str(y_data.get('ar_date', ''))
                            if old_br_dt == 'None': old_br_dt = ''
                            if old_ar_dt == 'None': old_ar_dt = ''
                            
                            new_br_by = str(row_new.get(f'{y} BR Paid By', '')).strip()
                            if new_br_by == '': new_br_by = 'Firm'
                            
                            raw_br_dt = to_date(row_new.get(f'{y} BR Date'))
                            new_br_dt = raw_br_dt.strftime('%Y-%m-%d') if raw_br_dt else ''
                            
                            raw_ar_dt = to_date(row_new.get(f'{y} AR Date'))
                            new_ar_dt = raw_ar_dt.strftime('%Y-%m-%d') if raw_ar_dt else ''
                            
                            if old_br_by != new_br_by: diff_list.append({"Company": en_name, "Field": f"{y} BR Paid By", "Old Value": old_br_by, "New Value": new_br_by})
                            if old_br_dt != new_br_dt: diff_list.append({"Company": en_name, "Field": f"{y} BR Date", "Old Value": old_br_dt.replace('-','/') if old_br_dt else '', "New Value": new_br_dt.replace('-','/') if new_br_dt else ''})
                            if old_ar_dt != new_ar_dt: diff_list.append({"Company": en_name, "Field": f"{y} AR Date", "Old Value": old_ar_dt.replace('-','/') if old_ar_dt else '', "New Value": new_ar_dt.replace('-','/') if new_ar_dt else ''})
                            
                    else:
                        diff_list.append({"Company": en_name, "Field": "NEW RECORD", "Old Value": "N/A", "New Value": "Will be added"})

                if diff_list or True: 
                    st.subheader("🔍 Review Changes")
                    if diff_list: st.table(pd.DataFrame(diff_list))
                    else: st.info("No changes detected in the file. Click Sync to proceed anyway.")
                    
                    if st.button("🚀 Confirm & Apply Changes", key="btn_final_sync_v166"):
                        up_df['compliance_records'] = "{}"
                        
                        for idx, row_new in up_df.iterrows():
                            base_dt = get_base_date(row_new)
                            inc_yr = base_dt.year if base_dt else None
                            
                            comp_dict = {}
                            for y in active_years:
                                y_str = str(y)
                                br_by = str(row_new.get(f'{y} BR Paid By', 'Firm')).strip()
                                if br_by == '': br_by = 'Firm'
                                
                                raw_br = to_date(row_new.get(f'{y} BR Date'))
                                raw_ar = to_date(row_new.get(f'{y} AR Date'))
                                
                                if inc_yr and y < inc_yr:
                                    br_by = 'N/A'
                                    raw_br = None
                                    raw_ar = None
                                elif inc_yr and y == inc_yr:
                                    raw_ar = None
                                
                                if br_by == 'N/A': raw_br = None
                                
                                comp_dict[y_str] = {
                                    "br_paid_by": br_by,
                                    "br_date": raw_br.strftime('%Y-%m-%d') if raw_br else None,
                                    "ar_date": raw_ar.strftime('%Y-%m-%d') if raw_ar else None
                                }
                            up_df.at[idx, 'compliance_records'] = json.dumps(comp_dict)
                        
                        up_df = up_df.drop(columns=[f"{y} BR Paid By" for y in active_years] + [f"{y} BR Date" for y in active_years] + [f"{y} AR Date" for y in active_years], errors='ignore')
                        combined_df = pd.concat([existing_df, up_df]).drop_duplicates(subset=['_anchor'], keep='last')
                        combined_df = combined_df.drop(columns=['_anchor'], errors='ignore')
                            
                        combined_df.to_sql('companies', engine, if_exists='replace', index=False)
                        st.success("✅ Sync Completed!"); st.balloons(); st.rerun()
        except Exception as e: st.error(f"Error: {e}")
